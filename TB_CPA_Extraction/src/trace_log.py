"""
trace_log.py  —  TB_CPA_Extraction v1.2
Persistent per-archive audit log (one row per ZIP archive, upsert on rerun).

Multi-PC safety: each PC writes to its own file:
    extraction_trace_log_{HOSTNAME}.xlsx
The dashboard merges all per-PC logs at generation time.
"""

import json
from pathlib import Path
import pandas as pd
from datetime import datetime
import re

# ── Column order for the Excel sheet ─────────────────────────────────────────
_COLUMNS = [
    "Run_timestamp",
    "PC_hostname",
    "ZIP_name",
    "ZIP_path",
    "To_copy",
    "Copied",
    "Corrupt",
    "Ignored",
    "Unknown",
    "Cell_IDs",             # comma-separated string
    "Corrupt_files_json",   # JSON string: {cellid: [filenames]}
    "Archive_moved",        # True | False | —
    "Status",               # Success | Partial | Failed
]


class ExtractionTraceLog:
    """
    Upsert-based trace log — one row per ZIP archive (keyed by ZIP_path).

    Upsert rule: if a row for this ZIP_path already exists it is overwritten
    (unlike harmonize v1.2, re-running extraction always reflects the latest state).
    """

    def __init__(self, log_path: Path, hostname: str = ""):
        self.log_path  = log_path
        self._hostname = hostname

        if log_path.exists():
            self._df = pd.read_excel(log_path, dtype=str)
            for col in _COLUMNS:
                if col not in self._df.columns:
                    self._df[col] = "—"
            self._df = self._df[_COLUMNS]
        else:
            self._df = pd.DataFrame(columns=_COLUMNS)

    # ── Public API ────────────────────────────────────────────────────────────

    def record_run(self, status_dict: dict, run_timestamp: str):
        """
        Upsert one row per ZIP archive from the pipeline status dict.

        status_dict keys are full paths to the ZIP archive.
        """
        for zip_path_str, entry in status_dict.items():
            self._upsert_row(zip_path_str, entry, run_timestamp)

    def save(self):
        """Write the trace log to Excel."""
        _write_excel(self._df, self.log_path)
        print(f"[TraceLog] Saved → {self.log_path.name}  ({len(self._df)} total rows)")

    @property
    def df(self) -> pd.DataFrame:
        return self._df.copy()

    # ── Private ───────────────────────────────────────────────────────────────

    def _upsert_row(self, zip_path_str: str, entry: dict, run_timestamp: str):
        to_copy  = len(entry.get("to_copy", {}).get("meta", {}))
        copied   = len(entry.get("copied_files_meta", {}))
        corrupt  = len(entry.get("corrupted", {}).get("names", []))
        ignored  = len(entry.get("to_ignore", {}).get("names", []))
        unknown  = len(entry.get("unknown", {}).get("names", []))

        # Cell IDs from copied files
        cell_ids = sorted({
            meta.get("cellid", "?")
            for meta in entry.get("copied_files_meta", {}).values()
            if meta.get("cellid")
        })

        # Corrupt files grouped by cell ID
        corrupt_by_cell = _group_corrupt_by_cellid(
            entry.get("corrupted", {}).get("names", []),
            entry.get("to_copy", {}).get("meta", {}),
        )

        # Archive moved?
        comp_meta = entry.get("compressed_file_meta", {})
        archive_moved = "True" if comp_meta.get("copied_to_Archived") else "False"

        # Overall status
        if corrupt == 0 and unknown == 0:
            status = "Success"
        elif copied > 0:
            status = "Partial"
        else:
            status = "Failed"

        row_data = {
            "Run_timestamp":    run_timestamp,
            "PC_hostname":      self._hostname,
            "ZIP_name":         Path(zip_path_str).name,
            "ZIP_path":         zip_path_str,
            "To_copy":          str(to_copy),
            "Copied":           str(copied),
            "Corrupt":          str(corrupt),
            "Ignored":          str(ignored),
            "Unknown":          str(unknown),
            "Cell_IDs":         ", ".join(cell_ids) if cell_ids else "—",
            "Corrupt_files_json": json.dumps(corrupt_by_cell) if corrupt_by_cell else "—",
            "Archive_moved":    archive_moved,
            "Status":           status,
        }

        existing_mask = self._df["ZIP_path"] == zip_path_str
        if existing_mask.any():
            idx = self._df[existing_mask].index[0]
            for k, v in row_data.items():
                self._df.at[idx, k] = v
        else:
            new_row = pd.DataFrame([row_data], columns=_COLUMNS)
            self._df = pd.concat([self._df, new_row], ignore_index=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _group_corrupt_by_cellid(corrupt_names: list, to_copy_meta: dict) -> dict:
    """
    Group corrupt filenames by cell ID.
    Uses the to_copy meta (which has cellid_prefix) to infer IDs; falls back to
    the same logic as extract_cellid_from_name in file_handling.py.
    """
    result: dict[str, list] = {}
    for name in corrupt_names:
        stem = Path(name).stem
        cellid = _infer_cellid(stem, to_copy_meta, name)
        result.setdefault(cellid, []).append(Path(name).name)
    return result


def _infer_cellid(stem: str, to_copy_meta: dict, original_name: str) -> str:
    """
    Try to find the cell ID for a corrupt file.
    First checks if the file appeared in to_copy_meta (before corruption was detected);
    if not, falls back to regex splitting on the stem.
    """
    for path_key, meta in to_copy_meta.items():
        if Path(path_key).name == Path(original_name).name:
            prefix = meta.get("cellid_prefix", "")
            if prefix and prefix in stem:
                splits = stem.split(prefix)
                if len(splits) > 1:
                    return prefix + re.split(r"[^a-zA-Z0-9]+", splits[1])[0]
            return meta.get("cellid", "unknown")
    return "unknown"


# ── Excel writer ─────────────────────────────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Extraction Log"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        STATUS_COLORS = {
            "Success": "C6EFCE",
            "Partial": "FFEB9C",
            "Failed":  "FFC7CE",
        }
        status_col_idx = list(df.columns).index("Status") + 1

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if pd.notna(value) else "—")
                cell.alignment = Alignment(vertical="center", wrap_text=False)
            status_val = str(row.get("Status", ""))
            if status_val in STATUS_COLORS:
                ws.cell(row=row_idx, column=status_col_idx).fill = PatternFill("solid", fgColor=STATUS_COLORS[status_val])

        COL_WIDTHS = {
            "Run_timestamp": 20, "PC_hostname": 16, "ZIP_name": 40, "ZIP_path": 70,
            "To_copy": 10, "Copied": 10, "Corrupt": 10, "Ignored": 10, "Unknown": 10,
            "Cell_IDs": 40, "Corrupt_files_json": 60, "Archive_moved": 16, "Status": 12,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 16)

        ws.freeze_panes = "A2"
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    except ImportError:
        csv_path = path.with_suffix(".csv")
        df.to_csv(csv_path, index=False)
        print(f"[TraceLog] openpyxl missing; saved as CSV → {csv_path.name}")
