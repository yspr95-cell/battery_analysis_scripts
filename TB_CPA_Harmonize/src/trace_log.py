"""
trace_log.py  —  TB_CPA_Harmonize v1.2
Persistent per-file audit log (one row per source file, upsert on rerun).

Multi-PC safety: each PC writes to its own file:
    harmonize_trace_log_{HOSTNAME}.xlsx
The dashboard merges all per-PC logs at generation time.
"""

from pathlib import Path
import pandas as pd
import os
from datetime import datetime

# ── Column order for the Excel sheet ────────────────────────────────────────
_COLUMNS = [
    "Run_timestamp",
    "PC_hostname",
    "Cell_ID",
    "File_name",
    "File_path",
    "File_size_KB",
    "Supplier",
    "Config_used",
    "Status",           # Harmonized | Skipped | Failed | No_config
    "Skip_reason",      # already_harmonized | no_config_match | empty_output | error | —
    "Error_message",
    "Harmonized_file_path",
    "Output_size_KB",   # size of the harmonized CSV at time of export
    "Row_count",
    "Date_harmonized",
    "Current_status",   # OK | Modified | Deleted | Not_applicable
]

_MODIFIED_THRESHOLD_KB = 10  # treat as Modified if size diff exceeds this


class TraceLog:
    """
    Upsert-based trace log — one row per source file (keyed by File_path).

    Upsert rules:
      - Status=Skipped + row already exists → do nothing (preserve previous status)
      - Status=Skipped + no row yet         → insert new row
      - Any other status + row exists       → update row in place
      - Any other status + no row           → insert new row
    """

    def __init__(self, log_path: Path, hostname: str = ""):
        self.log_path = log_path
        self._hostname = hostname
        if log_path.exists():
            self._df = pd.read_excel(log_path, dtype=str)
            # ensure all expected columns exist (handles older log files)
            for col in _COLUMNS:
                if col not in self._df.columns:
                    self._df[col] = "—"
            self._df = self._df[_COLUMNS]  # enforce column order
        else:
            self._df = pd.DataFrame(columns=_COLUMNS)

        self._changed = False  # track if anything was modified

    # ── Public API ────────────────────────────────────────────────────────────

    def record(
        self,
        run_timestamp: str,
        cell_id: str,
        file_path: Path,
        *,
        supplier: str = "—",
        config_used: str = "—",
        status: str,              # Harmonized | Skipped | Failed | No_config
        skip_reason: str = "—",
        error_message: str = "—",
        harmonized_file_path: Path = None,
        row_count=None,
    ):
        """
        Upsert one row for this file.
        If status=Skipped and the file already has a row, the row is left unchanged.
        """
        file_path_str = str(file_path)
        existing_mask = self._df["File_path"] == file_path_str
        row_exists = bool(existing_mask.any())

        # Skip-on-existing rule: 'Skipped' never overwrites a previous record
        if status == "Skipped" and row_exists:
            return

        file_size_kb = _get_size_kb(file_path)
        output_size_kb = "—"
        date_harmonized = "—"
        current_status = "Not_applicable"

        if harmonized_file_path is not None and Path(harmonized_file_path).exists():
            output_size_kb = _get_size_kb(Path(harmonized_file_path))
            date_harmonized = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            current_status = "OK"

        row_data = {
            "Run_timestamp":        run_timestamp,
            "PC_hostname":          self._hostname,
            "Cell_ID":              cell_id,
            "File_name":            file_path.name,
            "File_path":            file_path_str,
            "File_size_KB":         file_size_kb,
            "Supplier":             supplier,
            "Config_used":          config_used,
            "Status":               status,
            "Skip_reason":          skip_reason,
            "Error_message":        str(error_message)[:500] if error_message != "—" else "—",
            "Harmonized_file_path": str(harmonized_file_path) if harmonized_file_path else "—",
            "Output_size_KB":       output_size_kb,
            "Row_count":            str(row_count) if row_count is not None else "—",
            "Date_harmonized":      date_harmonized,
            "Current_status":       current_status,
        }

        if row_exists:
            idx = self._df[existing_mask].index[0]
            for k, v in row_data.items():
                self._df.at[idx, k] = v
        else:
            new_row = pd.DataFrame([row_data], columns=_COLUMNS)
            self._df = pd.concat([self._df, new_row], ignore_index=True)

        self._changed = True

    def update_current_status(self):
        """
        Refresh Current_status for ALL rows that were previously harmonized.
        Called before saving so the log always reflects real on-disk state.

        Rules:
          - Harmonized_file_path == '—'          → Not_applicable
          - File missing                          → Deleted
          - |current_size - stored_size| > 10 KB  → Modified
          - Otherwise                             → OK
        """
        for idx, row in self._df.iterrows():
            hp = row.get("Harmonized_file_path", "—")
            if hp == "—" or pd.isna(hp):
                self._df.at[idx, "Current_status"] = "Not_applicable"
                continue

            hp_path = Path(hp)
            if not hp_path.exists():
                self._df.at[idx, "Current_status"] = "Deleted"
                continue

            try:
                stored_kb = float(row.get("Output_size_KB", 0) or 0)
                current_kb = _get_size_kb(hp_path)
                if abs(current_kb - stored_kb) > _MODIFIED_THRESHOLD_KB:
                    self._df.at[idx, "Current_status"] = "Modified"
                else:
                    self._df.at[idx, "Current_status"] = "OK"
            except Exception:
                self._df.at[idx, "Current_status"] = "OK"

    def save(self):
        """Update Current_status for all rows, then write to Excel."""
        self.update_current_status()
        _write_excel(self._df, self.log_path)
        n = int(self._changed)
        print(f"[TraceLog] Saved → {self.log_path.name}  ({len(self._df)} total rows)")

    @property
    def df(self) -> pd.DataFrame:
        """Return the current in-memory DataFrame."""
        return self._df.copy()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_size_kb(path: Path) -> float:
    try:
        return round(path.stat().st_size / 1024, 2)
    except Exception:
        return 0.0


def _write_excel(df: pd.DataFrame, path: Path):
    """Write DataFrame to xlsx with basic formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Harmonize Log"

        # ── Header row ────────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Data rows ─────────────────────────────────────────────────────────
        STATUS_COLORS = {
            "Harmonized":    "C6EFCE",  # light green
            "Skipped":       "FFEB9C",  # light yellow
            "Failed":        "FFC7CE",  # light red
            "No_config":     "FFC7CE",
        }
        CURRENT_COLORS = {
            "OK":            "C6EFCE",
            "Modified":      "FFEB9C",
            "Deleted":       "FFC7CE",
            "Not_applicable": "F2F2F2",
        }

        status_col_idx = list(df.columns).index("Status") + 1
        current_col_idx = list(df.columns).index("Current_status") + 1

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if pd.notna(value) else "—")
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            # colour Status cell
            status_val = str(row.get("Status", ""))
            if status_val in STATUS_COLORS:
                ws.cell(row=row_idx, column=status_col_idx).fill = PatternFill("solid", fgColor=STATUS_COLORS[status_val])

            # colour Current_status cell
            cs_val = str(row.get("Current_status", ""))
            if cs_val in CURRENT_COLORS:
                ws.cell(row=row_idx, column=current_col_idx).fill = PatternFill("solid", fgColor=CURRENT_COLORS[cs_val])

        # ── Column widths ──────────────────────────────────────────────────────
        COL_WIDTHS = {
            "Run_timestamp": 20, "Cell_ID": 18, "File_name": 38, "File_path": 60,
            "File_size_KB": 14, "Supplier": 12, "Config_used": 20, "Status": 14,
            "Skip_reason": 22, "Error_message": 40, "Harmonized_file_path": 60,
            "Output_size_KB": 16, "Row_count": 12, "Date_harmonized": 20,
            "Current_status": 16,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 16)

        ws.freeze_panes = "A2"
        wb.save(path)
    except ImportError:
        # Fallback: plain CSV if openpyxl is somehow missing
        csv_path = path.with_suffix(".csv")
        df.to_csv(csv_path, index=False)
        print(f"[TraceLog] openpyxl not found; saved as CSV → {csv_path.name}")
