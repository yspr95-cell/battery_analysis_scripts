"""File loading utilities for Excel and CSV files.

Optimized for large files (5M+ rows):
- Header detection only reads first N rows
- Preview mode reads a limited number of rows
- Full loading uses pyarrow engine for CSV when available
- Separator/encoding detection reads only a small sample
"""

from pathlib import Path
import pandas as pd
import logging

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.tsv'}

# How many rows to read for header detection and preview
_HEADER_SCAN_ROWS = 50
_DEFAULT_PREVIEW_ROWS = 1000


def get_file_type(filepath: Path) -> str:
    """Return file type category: 'excel', 'csv', or 'unknown'."""
    ext = filepath.suffix.lower()
    if ext in ('.xlsx', '.xls'):
        return 'excel'
    elif ext in ('.csv', '.tsv'):
        return 'csv'
    return 'unknown'


def get_sheet_names(filepath: Path) -> list[str]:
    """Return list of sheet names for Excel files, or ['Sheet1'] for CSV."""
    file_type = get_file_type(filepath)
    if file_type == 'excel':
        xls = pd.ExcelFile(filepath)
        names = xls.sheet_names
        xls.close()
        return names
    elif file_type == 'csv':
        return ['Sheet1']
    return []


def get_file_size_mb(filepath: Path) -> float:
    """Return file size in MB."""
    return filepath.stat().st_size / (1024 * 1024)


def load_sheet(filepath: Path, sheet_name: str | None = None,
               header_row: int = 0, encoding: str = 'utf-8',
               separator: str | None = None,
               nrows: int | None = None) -> pd.DataFrame:
    """Load a single sheet/file into a DataFrame.

    Args:
        filepath: Path to file.
        sheet_name: Sheet name for Excel files (ignored for CSV).
        header_row: 0-indexed row number to use as column headers.
        encoding: Character encoding for CSV files.
        separator: Column separator for CSV files (auto-detected if None).
        nrows: Maximum number of data rows to read (None = all rows).

    Returns:
        DataFrame with data starting after the header row.
    """
    file_type = get_file_type(filepath)

    if file_type == 'excel':
        df = pd.read_excel(filepath, sheet_name=sheet_name,
                           header=header_row, nrows=nrows)
    elif file_type == 'csv':
        if separator is None:
            separator = _detect_separator(filepath, encoding)
        df = _read_csv_safe(filepath, header=header_row, encoding=encoding,
                            sep=separator, nrows=nrows)
    else:
        raise ValueError(f"Unsupported file type: {filepath.suffix}")

    # Clean column names: strip whitespace
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_preview(filepath: Path, sheet_name: str | None = None,
                 header_row: int = 0, encoding: str = 'utf-8',
                 separator: str | None = None,
                 max_rows: int = _DEFAULT_PREVIEW_ROWS) -> tuple[pd.DataFrame, int]:
    """Load only the first max_rows for preview, plus count total rows efficiently.

    Returns:
        (preview_df, total_row_count)
    """
    file_type = get_file_type(filepath)

    # Load preview subset
    preview_df = load_sheet(filepath, sheet_name=sheet_name,
                            header_row=header_row, encoding=encoding,
                            separator=separator, nrows=max_rows)

    # Count total rows without loading full data
    total_rows = _count_rows(filepath, file_type, sheet_name=sheet_name,
                             header_row=header_row, encoding=encoding,
                             separator=separator)

    return preview_df, total_rows


def detect_header_row(filepath: Path, sheet_name: str | None = None,
                      encoding: str = 'utf-8', separator: str | None = None,
                      max_scan_rows: int = 20) -> int:
    """Auto-detect the header row. Only reads first ~50 rows, not the full file.

    Heuristic: a header row has many columns filled with short text labels.
    Metadata rows typically have few filled columns or long text in one cell.

    Returns 0-indexed row number.
    """
    df_raw = _load_raw_sample(filepath, sheet_name=sheet_name,
                              encoding=encoding, separator=separator,
                              nrows=_HEADER_SCAN_ROWS)

    scan_limit = min(max_scan_rows, len(df_raw))
    total_cols = len(df_raw.columns)
    best_row = 0
    best_score = 0

    for i in range(scan_limit):
        row = df_raw.iloc[i]
        string_count = 0
        non_empty_count = 0
        for val in row:
            if pd.notna(val):
                non_empty_count += 1
                if isinstance(val, str):
                    val_stripped = val.strip()
                    if val_stripped and not _is_numeric_string(val_stripped):
                        string_count += 1
        fill_ratio = non_empty_count / total_cols if total_cols > 0 else 0
        score = string_count * fill_ratio
        if score > best_score:
            best_score = score
            best_row = i

    return best_row


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _load_raw_sample(filepath: Path, sheet_name: str | None = None,
                     encoding: str = 'utf-8', separator: str | None = None,
                     nrows: int = _HEADER_SCAN_ROWS) -> pd.DataFrame:
    """Load first N rows with no header interpretation. Fast for any file size."""
    file_type = get_file_type(filepath)

    if file_type == 'excel':
        df = pd.read_excel(filepath, sheet_name=sheet_name,
                           header=None, nrows=nrows)
    elif file_type == 'csv':
        if separator is None:
            separator = _detect_separator(filepath, encoding)
        df = _read_csv_safe(filepath, header=None, encoding=encoding,
                            sep=separator, nrows=nrows)
    else:
        raise ValueError(f"Unsupported file type: {filepath.suffix}")
    return df


def _read_csv_safe(filepath: Path, encoding: str = 'utf-8',
                   sep: str = ',', header: int | None = 0,
                   nrows: int | None = None) -> pd.DataFrame:
    """Read CSV with encoding fallback, comment line handling, and pyarrow when possible."""
    kwargs = dict(sep=sep, header=header, comment='#', on_bad_lines='skip')
    if nrows is not None:
        kwargs['nrows'] = nrows

    # Use pyarrow engine for full reads (much faster for large files)
    # pyarrow doesn't support nrows, comment, or on_bad_lines, so only use for full reads
    use_pyarrow = (nrows is None and sep == ',' and header is not None)

    if use_pyarrow:
        try:
            return pd.read_csv(filepath, sep=sep, header=header,
                               encoding=encoding, engine='pyarrow')
        except Exception:
            pass  # Fall through to default engine

    try:
        return pd.read_csv(filepath, encoding=encoding, **kwargs)
    except UnicodeDecodeError:
        return pd.read_csv(filepath, encoding='iso-8859-1', **kwargs)


def _count_rows(filepath: Path, file_type: str,
                sheet_name: str | None = None, header_row: int = 0,
                encoding: str = 'utf-8', separator: str | None = None) -> int:
    """Count total data rows efficiently without loading full data into memory."""
    if file_type == 'excel':
        # For Excel, we must load the sheet - openpyxl has no row-count-only mode
        # Use read_only mode via openpyxl for .xlsx
        if filepath.suffix.lower() == '.xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                # max_row includes empty trailing rows sometimes, but it's fast
                count = ws.max_row - (header_row + 1) if ws.max_row else 0
                wb.close()
                return max(count, 0)
            except Exception:
                pass
        # Fallback: load full sheet (slow but correct)
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row,
                           usecols=[0])  # Only 1 column to save memory
        return len(df)

    elif file_type == 'csv':
        # Fast line count: read raw bytes, count newlines
        count = 0
        try:
            with open(filepath, 'rb') as f:
                # Skip comment lines and count the rest
                for line in f:
                    if not line.startswith(b'#'):
                        count += 1
            # Subtract header rows (header_row + 1 lines are not data)
            return max(count - (header_row + 1), 0)
        except Exception:
            return 0

    return 0


def _detect_separator(filepath: Path, encoding: str = 'utf-8') -> str:
    """Detect CSV separator by reading the first few lines."""
    try:
        with open(filepath, 'r', encoding=encoding) as f:
            sample = f.read(4096)
    except UnicodeDecodeError:
        with open(filepath, 'r', encoding='iso-8859-1') as f:
            sample = f.read(4096)

    candidates = {',': 0, ';': 0, '\t': 0, '|': 0}
    lines = sample.split('\n')[:5]
    for line in lines:
        for sep in candidates:
            candidates[sep] += line.count(sep)

    best = max(candidates, key=candidates.get)
    return best if candidates[best] > 0 else ','


def detect_data_start_row(df: pd.DataFrame, max_scan: int = 10) -> int:
    """Detect how many leading rows in a loaded DataFrame are non-data (metadata).

    Heuristic: a real data row has at least some numeric or datetime values.
    Metadata rows (e.g. report headers) are mostly text or empty.

    Args:
        df: DataFrame already loaded with the correct header row.
        max_scan: Maximum number of rows to inspect.

    Returns:
        Index of the first row that looks like actual data (0 = no skip needed).
    """
    if df.empty:
        return 0

    scan_limit = min(max_scan, len(df))

    for i in range(scan_limit):
        row = df.iloc[i]
        numeric_count = 0
        non_empty_count = 0

        for val in row:
            # Skip NaN and empty strings
            if pd.isna(val):
                continue
            if isinstance(val, str) and val.strip() == '':
                continue
            non_empty_count += 1
            # Count numeric and datetime-like values
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                numeric_count += 1
            elif hasattr(val, 'year'):  # datetime / Timestamp
                numeric_count += 1

        if non_empty_count == 0:
            continue  # fully empty row – keep scanning

        numeric_fraction = numeric_count / non_empty_count
        # Real data row: ≥25% numeric cells OR at least 2 numeric columns
        if numeric_fraction >= 0.25 or numeric_count >= 2:
            return i

    return 0  # could not detect, assume no skip needed


def _is_numeric_string(s: str) -> bool:
    """Check if a string looks like a number."""
    try:
        float(s.replace(',', ''))
        return True
    except (ValueError, AttributeError):
        return False
